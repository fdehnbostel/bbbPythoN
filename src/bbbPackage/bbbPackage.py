import os
import csv

# joblib
import joblib

# numpy
import numpy as np

# pandas
import pandas as pd

# openpyxl
from openpyxl import Workbook

# rdkit
from rdkit import Chem
from rdkit.Chem import AllChem
from rdkit.Chem import rdForceFieldHelpers

# mordred
from mordred import Calculator, descriptors



#----TYPE CHECK----#
def IS_TYPE(desc_value,wBool = 1):
    """
    Checks if given descriptor value is float, int or, if specified, bool.
    """
    is_float = isinstance(desc_value,float)
    is_int =   isinstance(desc_value,int) or isinstance(desc_value,np.int64)
    if wBool == 1:
        is_bool = isinstance(desc_value,np.bool_) or isinstance(desc_value,bool)
        is_any = is_float or is_int or is_bool
        return(is_any)
    else:
        is_any = is_float or is_int
        return(is_any)

#----SMILES TO MOLS----#
def SMILES_TO_MOLS(smilesList):
    """
    Produces molecule objects from SMILES.
    """
    mols = []
    for smiles in smilesList:
         mols.append(Chem.MolFromSmiles(smiles))
    return(mols)

#----DATASET FETCHING----#

def DATA_FROM_CSV(filepath,smiles_pos=1,
                  id_pos=0,skip_first=False):
    """
    Retrieves smiles and ids from .csv file.
    """
    smiles = []
    ids = []
    with open(filepath,"r") as file:
        reader = csv.reader(file,delimiter=",")
        if skip_first:
            reader.next()
        for row in reader:
            smiles.append(row[smiles_pos])
            ids.append(row[id_pos])
    return(smiles,ids)
    
def MOLS_FROM_SDF(filepath,id_name="ID"):
    """
    Fetching molecules from .sdf files.
    """
    with Chem.SDMolSupplier(filepath) as suppl:
        mols = []
        ids = []
        for i,mol in enumerate(suppl):
            mols.append(mol)
            try:
                ids.append(mol.GetProp(id_name)) 
            except KeyError:
                ids.append("mol. {}".format(i))
    return(mols,ids)

#----FETCHING DESCRIPTOR NAMES AND BIN BOUNDARIES----#

def READ_BIN_LIMS(desc_dim,nBits,ks_dist):
    """
    Fetches the limits of the bins used for binary encoding of 
    MI-DSE descriptor values.
    """
    dir_path = os.path.dirname(os.path.realpath(__file__))
    ks_dict = {"eu": "euclidean", "ma": "mahalanobis"}
    filename = os.path.join(dir_path,"bin_lims",ks_dict.get(ks_dist,""),
                            "{}_{}-bits_{}-dist_bin_lims.csv".format(desc_dim,nBits,ks_dist))    
    name = []
    bin_lims = []
    with open(filename,"r") as file:
        reader = csv.reader(file,delimiter="\n")
        for row in reader:
            row_split = row[0].split(":")
            name.append(row_split[0])
            bin_lims.append((row_split[0],[float(val) for val in row_split[1].split(",")]))
    return(bin_lims)
    
def READ_CLS_DESC(desc_dim,check_midse=False,feat_sel=False):
    """
    Fetches names of structural key and MI-DSE descriptors.
    """
    if desc_dim == "2D+3D" and feat_sel == True:
        na_frac_diff_names_3d,midse_names_3d = READ_CLS_DESC("3D",check_midse=check_midse)
        na_frac_diff_names_2d,midse_names_2d = READ_CLS_DESC("2D",check_midse=check_midse)
        midse_names_2d3d = midse_names_2d + midse_names_3d
        return(na_frac_diff_names_2d,midse_names_2d3d)
    else:
        mi_dse_names = []
        dir_path = os.path.dirname(os.path.realpath(__file__))
        with open(os.path.join(dir_path,"desc_names","{}_mi_dse_descs.csv".format(desc_dim)),"r") as file:
            reader = csv.reader(file,delimiter = ";")
            row = next(reader)
            for tup in row:
                split_tup = tup.split(":")
                if check_midse:
                    mi_dse_names.append((split_tup[0],float(split_tup[1])))
                else:
                    mi_dse_names.append(split_tup[0])
        na_frac_diff_names = []
        filename = os.path.join(dir_path,"desc_names", "{}_na_descs.csv".format(desc_dim) )
        if os.stat(filename).st_size != 0:
            with open(filename,"r") as file:
                reader = csv.reader(file,delimiter = ";")
                row = next(reader)
                for name in row:
                    na_frac_diff_names.append(name)
        return(na_frac_diff_names,mi_dse_names)    

#----MOLECULAR COORDINATES AND DESCRIPTOR CALCULATION----#
def CALC_3D(mol):
    """
    Calculates the 3D coordinates of the given molecule using a Force
    Field of the rdkit package.
    """
    mol_3d = Chem.AddHs(mol)
    AllChem.EmbedMolecule(mol_3d,AllChem.ETKDG()) 
    if rdForceFieldHelpers.MMFFHasAllMoleculeParams(mol_3d):
        try:
            res = rdForceFieldHelpers.MMFFOptimizeMolecule(mol_3d,maxIters=200)
            if res == 1:                                          
                res = rdForceFieldHelpers.MMFFOptimizeMolecule(mol_3d,maxIters=10000) 
        except ValueError:
            mol_3d = "err"
    else:
        mol_3d = "err"
    return(mol_3d)
    
def PROD_3D_MOLS(mols):
    """
    Produces list of molecules with 3D coordinates.
    """
    mols_3d = []
    for i,mol in enumerate(mols):
        mols_3d.append(CALC_3D(mol))
    return(mols_3d)
         
def CALC_DESCS_3D(mols,midse_desc_names=[]):
    """
    Calculates 3D descriptors of given molecules.
    """
    mols = PROD_3D_MOLS(mols)
    calc = Calculator(descriptors, ignore_3D=False)
    if midse_desc_names:
        calc.descriptors = [desc for desc in calc.descriptors if str(desc) in midse_desc_names]
    mols = [mol for mol in mols if mol != "err"]
    # calculating 2D and 3D desriptors for given molecules
    df = calc.pandas(mols)
    return(df)    
        
def CALC_DESCS_2D(mols,midse_desc_names=[]):
    """
    Calculates 2D descriptors of given molecules.
    """
    calc = Calculator(descriptors, ignore_3D=True)
    if midse_desc_names:
        calc.descriptors = [desc for desc in calc.descriptors if str(desc) in midse_desc_names]
    # calculating 2D desriptors for given molecules
    df = calc.pandas(mols)
    return(df)

def GET_3D_DESC_NAMES():
    """
    Retrieves the names of 3D descriptors.
    """
    descs_3D, descs_2D = Calculator(descriptors, ignore_3D=False).descriptors, Calculator(descriptors, ignore_3D=True).descriptors
    descs_3D_names = [str(desc) for desc in descs_3D if desc not in descs_2D]
    return(descs_3D_names)

def DF_PROD_DESCS(desc_dim,mols,keep_desc_names=[]):
    """
    Creates pandas dataframe containing descriptors of specified 
    dimension.
    """
    if desc_dim == "2D":
        df = CALC_DESCS_2D(mols,midse_desc_names=keep_desc_names)
    elif desc_dim == "3D" or desc_dim == "2D+3D":
        df = CALC_DESCS_3D(mols,midse_desc_names=keep_desc_names)
        if desc_dim == "3D" and not keep_desc_names:
            # getting dataframe only consisting of 3D descriptors
            df = df[GET_3D_DESC_NAMES()]
    return(df)
    
#------Fingerprinting------#

def CHECK_BIN(val,bin_lims,nBins,fp=False):
    """
    Allocates a descriptor of a molecule to a bin based on its value.
    """
    if fp:
        max_ret_bin = nBins
        min_bin = 0
        ret_bin = 1
    else:
        max_ret_bin = nBins-1
        min_bin = 1
        ret_bin = 0
        
    if val > bin_lims[-1]:
        return(max_ret_bin)
    elif val <= bin_lims[min_bin]:
        return(0)
    else:
        for i in range(0,nBins-1):
            if bin_lims[i] < val <= bin_lims[i+1]:
                return(i+ret_bin)      
                              
def PROD_BIT_FP(desc_vals,na_names,bin_lims_list,nBits):
    """
    Produces fingerprints for given molecule. 
    """
    desc_fp = []
    if na_names:
        for name in na_names: 
            if not IS_TYPE(desc_vals[name]):
                desc_fp.append(0)
            else:
                desc_fp.append(1)
    if nBits != 1:
        for bin_lims in bin_lims_list:
            bin_ind = CHECK_BIN(desc_vals[bin_lims[0]],bin_lims[1],nBits,fp=True)
            for i in range(0,nBits):
                if i < bin_ind:
                    desc_fp.append(1)
                else:
                    desc_fp.append(0)
        fp = np.asarray(desc_fp)
    else:
        for bin_lims in bin_lims_list:
            if desc_vals[bin_lims[0]].iloc[0] >= bin_lims[1][0]:
                desc_fp.append(1)
            else:
                desc_fp.append(0)
        fp = np.asarray(desc_fp)         
    return(fp)

#-------MODEL VALIDATION-------#     

def VAL_PREDS(result,probas,act,names):
    """
    Evaluates model predictions.
    """
    # Calculation of correct and false predictions
    pos_probas = [x[1] for x in probas]
    right_preds, false_preds = [pred for ind,pred in enumerate(result) if pred == act[ind]], [pred for ind,pred in enumerate(result) if pred != act[ind]]
    true_pos, true_neg = sum(right_preds), len(right_preds) - sum(right_preds)
    false_pos, false_neg = sum(false_preds), len(false_preds) - sum(false_preds) 
    pos_act, neg_act = sum(act), len(act) - sum(act)
    # Calculation of positive and negative predictions
    pos_preds, neg_preds = sum(result), len(result) - sum(result)
    # Saving names of false positives and false negatives of each Kfold 
    f_pos = [(names[ind],act[ind]) for ind,pred in enumerate(result) if pred == 1 and act[ind] == 0] 
    f_neg = [(names[ind],act[ind]) for ind,pred in enumerate(result) if pred == 0 and act[ind] == 1]
    return(pos_probas, right_preds, false_preds, true_pos, true_neg, false_pos, false_neg, pos_act, neg_act, pos_preds, neg_preds, f_pos, f_neg)

def TEST_VAL(result,probas,test_act,test_names,ext_val=0,tp_tn=0):
    """
    Calculates performance measures.
    """
    # Calculation of performance measures.
    pos_probas, right_preds, false_preds, true_pos, true_neg, false_pos, false_neg, pos_act, neg_act, pos_preds, neg_preds, f_pos, f_neg = VAL_PREDS(result,probas,test_act,test_names)
    acc = (true_pos+true_neg)/len(test_act) 
    prec = true_pos/(true_pos+false_pos)
    f_meas = 2*(prec*acc)/(prec+acc) 
    # accuracy = sensitivity if no negatives are present in dataset
    f_pos_probas = [round(float(str(probas[test_names.index(tup[0])][1])[:5]),3) for tup in f_pos]
    f_neg_probas = [round(float(str(probas[test_names.index(tup[0])][0])[:5]),3) for tup in f_neg]
    f_pos_names = [tup[0] for tup in f_pos]
    f_neg_names = [tup[0] for tup in f_neg]
    t_pos_name_prob = [(test_names[ind],round(float(str(probas[ind][1])[:5]),3)) for ind,res in enumerate(result) if res == 1 and test_names[ind] not in f_pos_names]
    t_neg_name_prob = [(test_names[ind],round(float(str(probas[ind][0])[:5]),3)) for ind,res in enumerate(result) if res == 0 and test_names[ind] not in f_neg_names]
    return(acc,f_meas,f_pos,f_neg,f_pos_probas,f_neg_probas,t_pos_name_prob,t_neg_name_prob)
        
#-------VALIDATION OUTPUT-------#             

def PROD_FP_FN_OUT_STRING(max_name_len,pred_list,pred_type,fold=0,probs=[]):
    """
    Creates output string that is used to save False Positives and
    False Negatives.
    """
    outString = "{}: \nName:".format(pred_type)
    for j in range(0,max_name_len):
        outString = outString + " "
    outString = outString + "Activity:"
    if not probs:
        outString = outString + "\n"
        for tup in pred_list:
            outString = outString + tup[0]
            for k in range(0,5+(max_name_len-len(tup[0]))):
                outString = outString + " "
            outString = outString + str(tup[1]) + "\n"
    else:
        for k in range(0,max_name_len):
            outString = outString + " "
        outString = outString + "Probability:\n"    
        for ind,tup in enumerate(pred_list):
            outString = outString + tup[0]
            for l in range(0,5+(max_name_len-len(tup[0]))):
                outString = outString + " "
            outString = outString + str(tup[1])# + "\n"
            for m in range(0,8+max_name_len):
                outString = outString + " "
            outString = outString + str(probs[ind]) + "\n"
    return(outString+"\n")        

def PROD_FP_FN_CSV_PCK(f_pos,f_neg,f_pos_prob=[],f_neg_prob=[],
                   t_pos_prob=[],t_pos_name_prob=[],t_neg_name_prob=[],vali="EXT"):
    """
    Writes .csv file of False Positives and False Negatives of external validations.
    """
    if not os.path.exists("perf_table"):
        os.makedirs("perf_table")
        
    filename = os.path.join("perf_table","fp_fn.csv")
    with open(filename,"w") as file:
        outString = ""
        if vali == "KFOLD":
            max_name_len = max(max([len(max(fold, key=lambda tup: len(tup[0]))[0]) if fold else 0 for fold in f_pos]),max([len(max(fold, key=lambda tup: len(tup[0]))[0]) if fold else 0 for fold in f_neg]))
            for i,fold in enumerate(f_pos):
                outString = outString + "Fold {}:\n".format(i) 
                outString = outString + PROD_FP_FN_OUT_STRING(max_name_len,fold,"False Positives",fold=1)
                outString = outString + PROD_FP_FN_OUT_STRING(max_name_len,f_neg[i],"False Negatives",fold=1)
                outString = outString + "\n\n"
        elif vali == "TEST" or vali == "EXT":
            if f_pos:
                max_name_len_pos = len(max(f_pos, key=lambda tup: len(tup[0]))[0])
            else:
                max_name_len_pos = 0
            if f_neg:
                max_name_len_neg = len(max(f_neg, key=lambda tup: len(tup[0]))[0])
            else:
                max_name_len_neg = 0
            if t_pos_name_prob:
                max_name_len_t_pos = len(max(t_pos_name_prob, key=lambda tup: len(tup[0]))[0])
            else:
                max_name_len_t_pos = 0
            if t_neg_name_prob:
                max_name_len_t_neg = len(max(t_neg_name_prob, key=lambda tup: len(tup[0]))[0])
            else:
                max_name_len_t_neg = 0
            
            max_name_len = max(max_name_len_pos,max_name_len_neg,max_name_len_t_pos,max_name_len_t_neg)
            
            outString = outString + PROD_FP_FN_OUT_STRING(max_name_len,f_pos,"False Positives",probs=f_pos_prob)
            outString = outString + PROD_FP_FN_OUT_STRING(max_name_len,f_neg,"False Negatives",probs=f_neg_prob)
            
            outString = outString + PROD_FP_FN_OUT_STRING(max_name_len,[(tup[0],1) for tup in t_pos_name_prob],"True Positives",probs=[tup[1] for tup in t_pos_name_prob])
            outString = outString + PROD_FP_FN_OUT_STRING(max_name_len,[(tup[0],0) for tup in t_neg_name_prob],"True Negatives",probs=[tup[1] for tup in t_neg_name_prob])
            outString = outString + "\n\n"
        file.write(outString)

def PROD_EXT_PERF_TAB_PCK(acc,f_meas,f_neg,num_mols):
    """
    Creates .xlsx file of performance measures of external validation.
    """
    wb = Workbook()
    ws = wb.active
    title = ""
    ws.title = "Performance Table"
    ws["A1"] = "Accuracy"
    ws["B1"] = "F-Measure"
    ws["C1"] = "FN / #molecules"
    ws["A{}".format(2)] = "%.3f" % acc
    ws["B{}".format(2)] = "%.3f" % f_meas
    ws["C{}".format(2)] = str(len(f_neg)) + "/" + str(num_mols)
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
   
    filename = os.path.join("perf_table","pred_perf.xlsx")
    wb.save(filename)

#-------VALIDATION OUTPUT-------#             
    
def ProdFP(smiles="",filepath="",id_name="",smiles_pos=-1,id_pos=-1,skip_first=False):
    """
    Produces Fingerprints for given SMILES, .csv containing SMILES, or .sdf containing molecules.
    smiles_pos, and id_pos define the position of the SMILES string and ID in the .csv file.
    id_name defines the property name holding the IDs of the molecules.
    """
    # load input smiles, 
    if smiles:
        inp_mols = SMILES_TO_MOLS([smiles])
    if filepath:
        fe = filepath.split(".")[-1]
        if fe == "csv":
            smiles,ids = DATA_FROM_CSV(filepath,smiles_pos=smiles_pos,
                                              id_pos=id_pos,skip_first=skip_first)
            inp_mols = SMILES_TO_MOLS(smiles)
        elif fe == "sdf":
            inp_mols,ids = MOLS_FROM_SDF(filepath,id_name=id_name)
        else:
            print("Please specify valid filetype (.csv or .sdf)")
    
    # calculate descriptors
    print("Calculating Descriptors...", end="",flush=True) 
    na_names,mi_dse_names = READ_CLS_DESC("2D+3D")
    keep_desc_names = na_names + mi_dse_names
    inp_df  = DF_PROD_DESCS("2D+3D",inp_mols[:10],keep_desc_names=keep_desc_names)
    print("Done", end="\n")
    
    print("Removing Molecules Producing Desriptor Errors...", end="",flush=True) 
    drop_rows = []
    for i,row in inp_df.iterrows():
        for desc in mi_dse_names:
            if not IS_TYPE(row[desc]):
                drop_rows.append(i)
    inp_df.drop(drop_rows,inplace=True)
    print("Done", end="\n")

    print("Producing {} Fingerprints...".format("2D+3D"), end="",flush=True)
    # featch bin boundaries
    bin_lims = READ_BIN_LIMS("2D+3D",8,"ma")
    # calculate fingerprints
    inp_df["Fingerprints"] = inp_df.apply(PROD_BIT_FP,args=(na_names,bin_lims,8),axis=1)
    inp_df["ID"] = ids
    print("Done", end="\n")
    return(inp_df)
    
def BbbPred(fps,ids,act=[],filepath = "model",filename="bbbRf.sav",ret=False):
    # fetch model
    dir_path = os.path.dirname(os.path.realpath(__file__))
    bbbRf = joblib.load(os.path.join(dir_path,"model","bbbRf.sav"))
    # perform prediction
    result = bbbRf.predict(fps)
    probas = bbbRf.predict_proba(fps)
    # report
    if act:
        # in case activity is supplied, performance metrics are calculated
        acc,f_meas,f_pos,f_neg,f_ppr,f_npr,t_pnpr,t_npr = TEST_VAL(result,probas,act,ids)
        #        
        PROD_FP_FN_CSV_PCK(f_pos,f_neg,f_pos_prob=f_ppr,f_neg_prob=f_npr,
                              t_pos_name_prob=t_pnpr,t_neg_name_prob=t_npr)  
        #
        PROD_EXT_PERF_TAB_PCK(acc,f_meas,f_neg,len(act))
    else:
        if ret:
            return(ids,result,probas)
        else:
            for i, pred in enumerate(result):
                print("Molecule {} prediction: {} (Prob.: {})".format(ids[i],pred,probas[i]))