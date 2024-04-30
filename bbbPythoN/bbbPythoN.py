import os
import csv

# joblib
import joblib

# numpy
import numpy as np

# pandas
import pandas as pd

# openpyxl
from openpyxl import Workbook

# rdkit
from rdkit import Chem
from rdkit.Chem import AllChem
from rdkit.Chem import rdForceFieldHelpers

# mordred
from mordred import Calculator, descriptors

#----TYPE CHECK----#
def IsType(desc_value,wBool = 1):
    """
    Checks if given descriptor value is float, int or, if specified, bool.
    """
    is_float = isinstance(desc_value,float)
    is_int =   isinstance(desc_value,int) or isinstance(desc_value,np.int64)
    if wBool == 1:
        is_bool = isinstance(desc_value,np.bool_) or isinstance(desc_value,bool)
        is_any = is_float or is_int or is_bool
        return(is_any)
    else:
        is_any = is_float or is_int
        return(is_any)

#----SMILES TO MOLS----#
def CleanMol(mol):
    """
    from
    https://bitsilla.com/blog/2021/06/standardizing-a-molecule-using-rdkit/    
    and
    https://www.blopig.com/blog/2022/05/molecular-standardization/
    based on
    https://github.com/greglandrum/RSC_OpenScience_Standardization_202104/blob/main/MolStandardize%20pieces.ipynb
    
    -Remove ions
    -Remove salt
    -Explicit hydrogen or implicit hydrogen
    
    -Aromatize benzene ring or kekule form
    -Neutralize
    """
        
    # save molecular properties before preprocessing
    mol_props = [(name,mol.GetProp(name)) for name in mol.GetPropNames()]
    
    # removeHs, disconnect metal atoms, normalize and reionize the molecule
    mol = rdMolStandardize.Cleanup(mol) 
    
    # kekulize
    Chem.Kekulize(mol)
    
    # identify parent fragment and neutralize molecule
    uncharger = rdMolStandardize.Uncharger() 
    mol = uncharger.uncharge(rdMolStandardize.FragmentParent(mol))
    
    # reset the properties on molecule
    for tup in mol_props:
        mol.SetProp(tup[0],tup[1])

    return(mol)
    
def SmilesToMol(smiles):
    """
    Produces molecule objects from SMILES.
    """
    mol = Chem.MolFromSmiles(smiles)
    if mol is not None:
        mol = CleanMol(mol)
        return(mol)

#----DATASET FETCHING----#

def DataFromCSV(filepath,id_name,act_name,smiles_name):
    # read csv file     
    csv_df = pd.read_csv(filepath)
    # calculate molecule objects from smiles
    csv_df["Mol"] =  csv_df[smiles_name].apply(SmilesToMol)
    # drop None / na molecules
    csv_df.dropna(subset=[smiles_name,id_name,act_name,"Mol"],inplace=True)
    # retrieving relevant columns
    inp_mols = csv_df["Mol"].tolist()
    ids = csv_df[id_name].tolist()
    act = csv_df[act_name].tolist()
    return(inp_mols,ids,act)
       
def MolsFromSDF(filepath,id_name,act_name):
    """
    Fetching molecules from .sdf file.
    """
    with Chem.SDMolSupplier(filepath) as suppl:
        mols = []
        ids = []
        act = []
        for i,mol in enumerate(suppl):
            if mol is not None:
                mol =  CleanMol(mol)
                if mol is not None:
                    mols.append(mol)
                    try:
                        ids.append(mol.GetProp(id_name)) 
                    except KeyError:
                        ids.append("mol. {}".format(i))
                    if act_name:
                        try:
                            act.append(int(mol.GetProp(act_name))) 
                        except KeyError:
                            print("Property holding activity not available!")
                            act.append(None)
                else:
                    print("Cleaning of Molecule with index {} resulted in None!".format(i))
            else:
                print("Molecule with index {} is None!".format(i))
    return(mols,ids,act) 

#----FETCHING DESCRIPTOR NAMES AND BIN BOUNDARIES----#

def ReadBinBounds(use_bal):
    """
    Fetches the limits of the bins used for binary encoding of 
    MI-DSE descriptor values.
    """
    dir_path = os.path.dirname(os.path.realpath(__file__))
    if use_bal:
        fn_suffix = "bal"
    else:
        fn_suffix = "imba"
    filename = os.path.join(dir_path,"bin_bounds","bin_bounds_{}.csv".format(fn_suffix))    
    name = []
    bin_bounds = []
    with open(filename,"r") as file:
        reader = csv.reader(file,delimiter="\n")
        for row in reader:
            row_split = row[0].split(",")
            name.append(row_split[0])
            bin_bounds.append((row_split[0],[float(val) for val in row_split[1:]]))
    return(bin_bounds)
    
def ReadClassDesc():
    """
    Fetches names of structural key and MI-DSE descriptors.
    """  
    mi_dse_names = []
    dir_path = os.path.dirname(os.path.realpath(__file__))
    with open(os.path.join(dir_path,"desc_names","mi-dse_descs.csv"),"r") as file:
        reader = csv.reader(file,delimiter = "\n")
        for row in reader:
            row_split = row[0].split(",")
            mi_dse_names.append(row_split[0])
        
    na_frac_diff_names = []
    filename = os.path.join(dir_path,"desc_names", "na_descs.csv")
    if os.stat(filename).st_size != 0:
        with open(filename,"r") as file:
            reader = csv.reader(file,delimiter = "\n")
            for row in reader:
                na_frac_diff_names.append(row[0])

    return(na_frac_diff_names,mi_dse_names)    

#----MOLECULAR COORDINATES AND DESCRIPTOR CALCULATION----#
def Calc3D(mol):
    """
    Calculates the 3D coordinates of the given molecule using a Force
    Field of the rdkit package.
    """
    mol_3d = Chem.AddHs(mol)
    AllChem.EmbedMolecule(mol_3d,AllChem.ETKDG()) 
    if rdForceFieldHelpers.MMFFHasAllMoleculeParams(mol_3d):
        try:
            res = rdForceFieldHelpers.MMFFOptimizeMolecule(mol_3d,maxIters=200)
            if res == 1:                                          
                res = rdForceFieldHelpers.MMFFOptimizeMolecule(mol_3d,maxIters=1000) 
        except ValueError:
            mol_3d = None
    else:
        mol_3d = None
    return(mol_3d)
    
def Prod3DMols(mols):
    """
    Produces list of molecules with 3D coordinates.
    """
    mols_3d = []
    for i,mol in enumerate(mols):
        mols_3d.append(Calc3D(mol))
    return(mols_3d)
         
def CalcDesc3D(mols,act=[],ids=[],midse_desc_names=[]):
    """
    Calculates 3D descriptors of given molecules.
    """
    mols = Prod3DMols(mols)
    calc = Calculator(descriptors, ignore_3D=False)
    if midse_desc_names:
        calc.descriptors = [desc for desc in calc.descriptors if str(desc) in midse_desc_names]
    if act:
        act = [act[i] for i,mol in enumerate(mols) if mol is not None]
    if ids:
        ids = [ids[i] for i,mol in enumerate(mols) if mol is not None]
    mols = [mol for mol in mols if mol is not None]
    # calculating 2D and 3D desriptors for given molecules
    df = calc.pandas(mols)
    if act:
        df["Act"] = act
    if ids:
        df["ID"] = ids
    return(df)    
        
def CalcDesc2D(mols,midse_desc_names=[]):
    """
    Calculates 2D descriptors of given molecules.
    """
    calc = Calculator(descriptors, ignore_3D=True)
    if midse_desc_names:
        calc.descriptors = [desc for desc in calc.descriptors if str(desc) in midse_desc_names]
    # calculating 2D desriptors for given molecules
    df = calc.pandas(mols)
    return(df)

def Get3DDescNames():
    """
    Retrieves the names of 3D descriptors.
    """
    descs_3D, descs_2D = Calculator(descriptors, ignore_3D=False).descriptors, \
                         Calculator(descriptors, ignore_3D=True).descriptors
    descs_3D_names = [str(desc) for desc in descs_3D if desc not in descs_2D]
    return(descs_3D_names)

def ProdDescDF(desc_dim,mols,act=[],ids=[],keep_desc_names=[]):
    """
    Creates pandas dataframe containing descriptors of specified 
    dimension.
    """
    if desc_dim == "2D":
        df = CalcDesc2D(mols,midse_desc_names=keep_desc_names)
        if act:
            df["Act"] = act
        if ids:
            df["ID"] = ids
    elif desc_dim == "3D" or desc_dim == "2D+3D":
        df = CalcDesc3D(mols,act=act,ids=ids,midse_desc_names=keep_desc_names)
        if desc_dim == "3D" and not keep_desc_names:
            # getting dataframe only consisting of 3D descriptors
            df = df[Get3DDescNames()]
    return(df)
    
#------Fingerprinting------#

def CheckBin(val,bin_bounds,nBins):
    """
    Allocates a descriptor of a molecule to a bin based on its value.
    """
    max_ret_bin = nBins
    min_bin = 0
    ret_bin = 1
    if val > bin_bounds[-1]:
        return(max_ret_bin)
    elif val <= bin_bounds[min_bin]:
        return(0)
    else:
        for i in range(0,nBins-1):
            if bin_bounds[i] < val <= bin_bounds[i+1]:
                return(i+ret_bin)      
                              
def ProdBitFP(desc_vals,na_names,bin_bounds_list,nBits):
    """
    Produces fingerprints for given molecule. 
    """
    desc_fp = []
    if na_names:
        for name in na_names: 
            if not IsType(desc_vals[name]):
                desc_fp.append(0)
            else:
                desc_fp.append(1)
    if nBits != 1:
        for bin_bounds in bin_bounds_list:
            bin_ind = CheckBin(desc_vals[bin_bounds[0]],bin_bounds[1],nBits)
            for i in range(0,nBits):
                if i < bin_ind:
                    desc_fp.append(1)
                else:
                    desc_fp.append(0)
        fp = np.asarray(desc_fp)
    else:
        for bin_bounds in bin_bounds_list:
            if desc_vals[bin_bounds[0]].iloc[0] >= bin_bounds[1][0]:
                desc_fp.append(1)
            else:
                desc_fp.append(0)
        fp = np.asarray(desc_fp)         
    return(fp)

#-------MODEL VALIDATION-------#     

def MdlVal(result,probas,act,names):
    """
    Evaluates model predictions.
    """
    # Calculation of correct and false predictions
    pos_probas = [x[1] for x in probas]
    right_preds, false_preds = [pred for ind,pred in enumerate(result) if pred == act[ind]], \
                               [pred for ind,pred in enumerate(result) if pred != act[ind]]
    true_pos, true_neg = sum(right_preds), len(right_preds) - sum(right_preds)
    false_pos, false_neg = sum(false_preds), len(false_preds) - sum(false_preds)
    pos_act, neg_act = sum(act), len(act) - sum(act)
    # Calculation of positive and negative predictions
    pos_preds, neg_preds = sum(result), len(result) - sum(result)
    # Saving names of false positives and false negatives 
    f_pos = [(names[ind],act[ind]) for ind,pred in enumerate(result) if pred == 1 and act[ind] == 0] 
    f_neg = [(names[ind],act[ind]) for ind,pred in enumerate(result) if pred == 0 and act[ind] == 1]
    return(pos_probas, right_preds, false_preds, true_pos, true_neg, false_pos, 
           false_neg, pos_act, neg_act, pos_preds, neg_preds, f_pos, f_neg)

def TestVal(result,probas,test_act,test_names,ext_val=0,tp_tn=0):
    """
    Calculates performance measures.
    """
    # Calculation of performance measures.
    pos_probas, right_preds, false_preds, true_pos, true_neg, false_pos, \
    false_neg, pos_act, neg_act, pos_preds, neg_preds, f_pos, f_neg = MdlVal(result,probas,test_act,test_names)
    acc = (true_pos+true_neg)/len(test_act) 
    sens = true_pos / (true_pos + false_neg)
    spec = true_neg / (true_neg + false_pos)
    prec = true_pos/(true_pos+false_pos)
    f_meas = 2*(prec*acc)/(prec+acc) 
    # accuracy = sensitivity if no negatives are present in dataset
    f_pos_probas = [round(float(str(probas[test_names.index(tup[0])][1])[:5]),3) for tup in f_pos]
    f_neg_probas = [round(float(str(probas[test_names.index(tup[0])][0])[:5]),3) for tup in f_neg]
    f_pos_names = [tup[0] for tup in f_pos]
    f_neg_names = [tup[0] for tup in f_neg]
    t_pos_name_prob = [(test_names[ind],round(float(str(probas[ind][1])[:5]),3)) for ind,res in enumerate(result) if 
                        res == 1 and test_names[ind] not in f_pos_names]
    t_neg_name_prob = [(test_names[ind],round(float(str(probas[ind][0])[:5]),3)) for ind,res in enumerate(result) if 
                        res == 0 and test_names[ind] not in f_neg_names]
    return(acc,sens,spec,f_meas,f_pos,f_neg,f_pos_probas,f_neg_probas,t_pos_name_prob,t_neg_name_prob)
        
#-------VALIDATION OUTPUT-------#    

def MahDist(x,y,cov_mat_inv):
    """
    Calculates the Mahalanobis Distance between the fingerprints of two molecules.
    """
    # creation of matrix epq which consists of differences between single descriptors of molecule x and y  
    epq = np.asarray([x_desc - y[ind] for ind,x_desc in enumerate(x)])
    # calculating mahalanobis distance between molecules x and y
    md = np.dot(cov_mat_inv,epq)
    md = np.dot(np.transpose(epq),md)
    md = np.sqrt(md)
    md = round(md,3)
    return(md)	         

def CalcMahDists(inp_df,use_bal):
    """
    Calculates Mahalanobis Distance for every input molecule.
    """
    dir_path = os.path.dirname(os.path.realpath(__file__))
    if use_bal:
        fn_suffix = "bal"
    else:
        fn_suffix = "imba"
    # fetching Mahalanobis distance information
    train_mv,train_cmi,train_ad = joblib.load(os.path.join(dir_path,"mah_dist_info",
                                             "train_mah-dist_info_2D+3D_ma_{}".format(fn_suffix)))
                                      
    train_amv,train_acmi,train_aad = joblib.load(os.path.join(dir_path,"mah_dist_info",
                                             "train_act_mah-dist_info_2D+3D_ma_{}".format(fn_suffix)))
    
    # fetching names of descriptors used in the MI-DSE fingerprint
    na_names,mi_dse_names = ReadClassDesc()
    descs_df = inp_df[mi_dse_names]
    
    # calculating distances 
    mah_dists = []
    mah_dists_act = []
    for i in range(descs_df.shape[0]):
        desc_row = descs_df.iloc[i].values.tolist()
        mah_dists.append(MahDist(desc_row,train_mv,train_cmi))
        mah_dists_act.append(MahDist(desc_row,train_amv,train_acmi))

    return(mah_dists,mah_dists_act)

def ProdFpFnOutString(max_name_len,pred_list,pred_type,fold=0,probs=[]):
    """
    Creates output string that is used to save False Positives and
    False Negatives.
    """
    outString = "{}: \nName:".format(pred_type)
    for j in range(0,max_name_len):
        outString = outString + " "
    outString = outString + "Activity:"
    if not probs:
        outString = outString + "\n"
        for tup in pred_list:
            outString = outString + tup[0]
            for k in range(0,5+(max_name_len-len(tup[0]))):
                outString = outString + " "
            outString = outString + str(tup[1]) + "\n"
    else:
        for k in range(0,max_name_len):
            outString = outString + " "
        outString = outString + "Probability:\n"    
        for ind,tup in enumerate(pred_list):
            outString = outString + tup[0]
            for l in range(0,5+(max_name_len-len(tup[0]))):
                outString = outString + " "
            outString = outString + str(tup[1])# + "\n"
            for m in range(0,8+max_name_len):
                outString = outString + " "
            outString = outString + str(probs[ind]) + "\n"
    return(outString+"\n")        

def ProdFpFnCSV(use_bal,f_pos,f_neg,f_pos_prob=[],f_neg_prob=[],
                t_pos_prob=[],t_pos_name_prob=[],t_neg_name_prob=[],vali="EXT"):
    """
    Writes .csv file of False Positives and False Negatives of external validations.
    """
    if not os.path.exists("perf_table"):
        os.makedirs("perf_table")
    if use_bal:
        fn_suffix = "bal"
    else:
        fn_suffix = "imba"
        
    filename = os.path.join("perf_table","fp_fn_{}.csv".format(fn_suffix))
    with open(filename,"w") as file:
        outString = ""
        if vali == "KFOLD":
            max_name_len = max(max([len(max(fold, key=lambda tup: len(tup[0]))[0]) if fold else 0 for fold in f_pos]), \
                           max([len(max(fold, key=lambda tup: len(tup[0]))[0]) if fold else 0 for fold in f_neg]))
            for i,fold in enumerate(f_pos):
                outString = outString + "Fold {}:\n".format(i) 
                outString = outString + ProdFpFnOutString(max_name_len,fold,"False Positives",fold=1)
                outString = outString + ProdFpFnOutString(max_name_len,f_neg[i],"False Negatives",fold=1)
                outString = outString + "\n\n"
        elif vali == "TEST" or vali == "EXT":
            if f_pos:
                max_name_len_pos = len(max(f_pos, key=lambda tup: len(tup[0]))[0])
            else:
                max_name_len_pos = 0
            if f_neg:
                max_name_len_neg = len(max(f_neg, key=lambda tup: len(tup[0]))[0])
            else:
                max_name_len_neg = 0
            if t_pos_name_prob:
                max_name_len_t_pos = len(max(t_pos_name_prob, key=lambda tup: len(tup[0]))[0])
            else:
                max_name_len_t_pos = 0
            if t_neg_name_prob:
                max_name_len_t_neg = len(max(t_neg_name_prob, key=lambda tup: len(tup[0]))[0])
            else:
                max_name_len_t_neg = 0
            
            max_name_len = max(max_name_len_pos,max_name_len_neg,max_name_len_t_pos,max_name_len_t_neg)
            
            outString = outString + ProdFpFnOutString(max_name_len,f_pos,"False Positives",probs=f_pos_prob)
            outString = outString + ProdFpFnOutString(max_name_len,f_neg,"False Negatives",probs=f_neg_prob)
            
            outString = outString + ProdFpFnOutString(max_name_len,[(tup[0],1) for tup in t_pos_name_prob],
                                                      "True Positives",probs=[tup[1] for tup in t_pos_name_prob])
            outString = outString + ProdFpFnOutString(max_name_len,[(tup[0],0) for tup in t_neg_name_prob],
                                                      "True Negatives",probs=[tup[1] for tup in t_neg_name_prob])
            outString = outString + "\n\n"
        file.write(outString)

def ProdExtPerfTab(use_bal,acc,sens,spec,f_meas,f_neg,f_pos,num_mols):
    """
    Creates .xlsx file of performance measures of external validation.
    """
    wb = Workbook()
    ws = wb.active
    title = ""
    ws.title = "Performance Table"
    ws["A1"] = "Accuracy"
    ws["B1"] = "Sensitivity"
    ws["C1"] = "Specificity"
    ws["D1"] = "F-Measure"
    ws["E1"] = "FN / #molecules"
    ws["A{}".format(2)] = "%.3f" % acc
    ws["B{}".format(2)] = "%.3f" % sens
    ws["C{}".format(2)] = "%.3f" % spec
    ws["D{}".format(2)] = "%.3f" % f_meas
    ws["E{}".format(2)] = str(len(f_neg)) + "/" + str(num_mols)
    ws["F{}".format(2)] = str(len(f_pos)) + "/" + str(num_mols)
    
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    
    if use_bal:
        fn_suffix = "bal"
    else:
        fn_suffix = "imba"
    filename = os.path.join("perf_table","pred_perf_{}.xlsx".format(fn_suffix))
    wb.save(filename)

#-------VALIDATION OUTPUT-------#             

def LoadInput(smiles,filepath,id_name,act_name,smiles_name):
    # load input smiles, 
    if smiles:
        ids = [0]
        act = [None]
        inp_mols = [SmilesToMol(smiles)]
    if filepath:
        fn, fe = os.path.splitext(filepath)
        if fe == ".csv":
            inp_mols,ids,act = DataFromCSV(filepath,id_name,act_name,smiles_name)
        elif fe == ".sdf":
            inp_mols,ids,act = MolsFromSDF(filepath,id_name,act_name=act_name)
        else:
            print("Please use valid filetype (.csv or .sdf)")
    return(inp_mols,ids,act)
    
    
def RemErrorMols(inp_df,mi_dse_names,na_names):
    drop_rows = []
    for i,row in inp_df.iterrows():
        drop = False
        for desc in mi_dse_names:
            # if descriptor value is neither float, int, or boolean
            if not IsType(row[desc]):
                # molecule having missing descriptor value is dropped
                drop = True
                break
        if drop:
            drop_rows.append(i)
            print("Removing Molecule {} ".format(row["ID"]))
        for desc in na_names:
            if isinstance(row[desc],float):
                inp_df.iloc[i, inp_df.columns.get_loc(desc)] = 1
            else:
                inp_df.iloc[i, inp_df.columns.get_loc(desc)] = 0
    inp_df.drop(drop_rows,inplace=True)

    
def ProdFP(smiles="",filepath="",id_name="",act_name="",
           smiles_name="",use_bal=False):
    """
    Produces Fingerprints for given SMILES, .csv containing SMILES, or .sdf containing molecules.
    smiles_pos, and id_pos define the position of the SMILES string and ID in the .csv file.
    id_name defines the property name holding the IDs of the molecules.
    """
    
    # loading dataset  
    inp_mols,ids,act = LoadInput(smiles,filepath,id_name,act_name,smiles_name)
           
    # calculate descriptors
    print("Calculating Descriptors...", end="",flush=True) 
    na_names,mi_dse_names = ReadClassDesc()
    keep_desc_names = na_names + mi_dse_names
    inp_df = ProdDescDF("2D+3D",inp_mols,act=act,ids=ids,keep_desc_names=keep_desc_names)
    print("Done", end="\n")
    
    # removing molecules with faulty descriptors
    print("Removing Molecules Producing Desriptor Errors...", end="",flush=True) 
    RemErrorMols(inp_df,mi_dse_names,na_names)
    print("Done", end="\n")

    
    print("Producing {} Fingerprints...".format("2D+3D"), end="",flush=True)
    # featch bin boundaries
    bin_bounds = ReadBinBounds(use_bal)
    # calculate fingerprints
    inp_df["Fingerprint"] = inp_df.apply(ProdBitFP,args=(na_names,bin_bounds,8),axis=1)
    print("Done", end="\n")
    return(inp_df)
    
def BbbPred(inp_df,act=False,ret=False,use_bal=False):
    """
    Performs predictions for given fingerprints. 
    In case activities are supplied a prediction evaluation is performed
    and its results saved to a .csv and an .xlsx file.
    """
    # fetch model
    dir_path = os.path.dirname(os.path.realpath(__file__))
    if use_bal:
        fn_suffix = "bal"
    else:
        fn_suffix = "imba"
    bbbRf = joblib.load(os.path.join(dir_path,"model","bbbRf_{}.sav".format(fn_suffix)))
    
    # fetching fingerprints, ids from dataframe
    fps = inp_df["Fingerprint"].tolist()
    ids = inp_df["ID"].tolist()
   
    if act:
        act = inp_df["Act"].tolist()
    
    # perform prediction
    result = bbbRf.predict(fps)
    probas = bbbRf.predict_proba(fps)
    
    # calculating Mahalanobis Distance to training center of query molecules
    mah_dists,mah_dists_act = CalcMahDists(inp_df,use_bal) 
    
    if act and (inp_df.shape[0] > 1):
        # in case activity is supplied, performance metrics are calculated
        acc,sens,spec,f_meas,f_pos,f_neg,f_ppr,f_npr,t_pnpr,t_npr = TestVal(result,probas,act,ids)
        # produces .csv file listing True Positives, True Negatives, False Positives, 
        # and False Negatives       
        ProdFpFnCSV(use_bal,f_pos,f_neg,f_pos_prob=f_ppr,f_neg_prob=f_npr,
                              t_pos_name_prob=t_pnpr,t_neg_name_prob=t_npr)  
        # produces .xlsx file of performance metrics
        ProdExtPerfTab(use_bal,acc,sens,spec,f_meas,f_neg,f_pos,len(act))
    else:
        act = [None for id in ids]
    if ret:            
        return(ids,result,probas,act)
    else:
        for i, pred in enumerate(result):
            print("""Molecule {} prediction: {} (Prob.: {}); Activity: {}
                 Distance to Training Set Center: {}
                 Distance to Training Set Actives Center: {}\n""".format(ids[i],pred,
                   probas[i],act[i],mah_dists[i],mah_dists_act[i]))
